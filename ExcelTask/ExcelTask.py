# Importing the libraries we need
import os
import random
from datetime import datetime
from openpyxl import Workbook
from faker import Faker
import smtplib
import getpass
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# SMTP settings
HOST = "smtp-mail.outlook.com"
PORT = 587
FROM_EMAIL = "221124@astanait.edu.kz"
TO_EMAIL = "tgrey01@mail.ru"
PASSWORD = getpass.getpass("Enter password: ")

# Function to generate Excel file
def generation():
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"
        
    # Generation random names
    fake = Faker()
    names = [fake.name() for i in range(10)]
        
    # Give names to the columns
    ws.cell(row=1, column=1, value="Names")
    ws.cell(row=1, column=2, value="Date")
    ws.cell(row=1, column=3, value="Time")

    # Filling data in a table
    for row in range(2, len(names) + 2):
        ws.cell(row=row, column=1, value=names[row - 2])
        ws.cell(row=row, column=2, value=datetime.now().date())
        ws.cell(row=row, column=3, value=datetime.now().time())
        
    # File name generation
    file_name = f"{random.choice(names)}_{datetime.now().date()}_{random.randint(100, 999)}.xlsx"
        
    # The path to save the file
    save_path = os.path.join(os.path.expanduser('~'), 'Documents', 'skcu', file_name)
        
    # File saving
    wb.save(save_path)
    print(f"The Excel file was successfully created and saved as {file_name} in the {save_path} folder")
        
    return save_path

# Calling the function to generate Excel file
file_path = generation()

# SMTP connection and sending the Excel file
if file_path:
        smtp = smtplib.SMTP(HOST, PORT)
        smtp.starttls()
        smtp.login(FROM_EMAIL, PASSWORD)

        msg = MIMEMultipart()
        msg['From'] = FROM_EMAIL
        msg['To'] = TO_EMAIL
        msg['Subject'] = "Excel File"

        part = MIMEBase('application', "octet-stream")
        with open(file_path, 'rb') as file:
            part.set_payload(file.read())

        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(file_path))
        msg.attach(part)

        smtp.sendmail(FROM_EMAIL, TO_EMAIL, msg.as_string())

        print("Excel file sent successfully!")
else:
    print("Unable to send email as file was not generated.")
